from __future__ import annotations

import csv
import io
import re
import secrets
from datetime import datetime

from django.contrib import messages
from django.contrib.auth import login, logout
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from django_ratelimit.decorators import ratelimit
from django.utils import timezone

from .forms import LoginForm, ProjectUserCreateForm, RegisterForm, SchoolRegistrationForm, SchoolReviewForm, UserUpdateForm
from .models import CustomUser, School


def _is_project_level_user(user) -> bool:
    return bool(
        getattr(user, "is_authenticated", False)
        and (user.is_superuser or getattr(user, "is_project_admin", False))
    )


def _can_manage_users(user) -> bool:
    return bool(
        user.is_authenticated
        and not getattr(user, "is_guest", False)
        and (
            user.is_superuser
            or getattr(user, "is_project_admin", False)
            or getattr(user, "is_admin", False)
            or getattr(user, "can_users", False)
        )
    )


_TRANS_MAP = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "zh", "з": "z",
    "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o", "п": "p", "р": "r",
    "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "sch",
    "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
}


def _translit_login_legacy1(full_name: str) -> str:
    text = (full_name or "").strip().lower().replace("ё", "е")
    out = []
    for ch in text:
        if ch in _TRANS_MAP:
            out.append(_TRANS_MAP[ch])
        elif "a" <= ch <= "z" or "0" <= ch <= "9":
            out.append(ch)
        else:
            out.append("_")
    base = re.sub(r"_+", "_", "".join(out)).strip("_")
    return base or "teacher"


_CYR_TRANS_MAP = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e", "ж": "zh", "з": "z",
    "и": "i", "й": "y", "к": "k", "л": "l", "м": "m", "н": "n", "о": "o", "п": "p", "р": "r",
    "с": "s", "т": "t", "у": "u", "ф": "f", "х": "h", "ц": "ts", "ч": "ch", "ш": "sh", "щ": "sch",
    "ъ": "", "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya",
}


def _translit_login_legacy2(full_name: str) -> str:
    text = (full_name or "").strip().lower().replace("ё", "е")
    out = []
    for ch in text:
        if ch in _CYR_TRANS_MAP:
            out.append(_CYR_TRANS_MAP[ch])
        elif "a" <= ch <= "z" or "0" <= ch <= "9":
            out.append(ch)
        else:
            out.append("_")
    base = re.sub(r"_+", "_", "".join(out)).strip("_")
    return base or "teacher"


_SAFE_CYR_TRANS_MAP = {
    "\u0430": "a", "\u0431": "b", "\u0432": "v", "\u0433": "g", "\u0434": "d", "\u0435": "e", "\u0451": "e", "\u0436": "zh", "\u0437": "z",
    "\u0438": "i", "\u0439": "y", "\u043a": "k", "\u043b": "l", "\u043c": "m", "\u043d": "n", "\u043e": "o", "\u043f": "p", "\u0440": "r",
    "\u0441": "s", "\u0442": "t", "\u0443": "u", "\u0444": "f", "\u0445": "h", "\u0446": "ts", "\u0447": "ch", "\u0448": "sh", "\u0449": "sch",
    "\u044a": "", "\u044b": "y", "\u044c": "", "\u044d": "e", "\u044e": "yu", "\u044f": "ya",
}


def _translit_login(full_name: str) -> str:
    text = (full_name or "").strip().lower().replace("\u0451", "\u0435")
    out = []
    for ch in text:
        if ch in _SAFE_CYR_TRANS_MAP:
            out.append(_SAFE_CYR_TRANS_MAP[ch])
        elif "a" <= ch <= "z" or "0" <= ch <= "9":
            out.append(ch)
        else:
            out.append("_")
    base = re.sub(r"_+", "_", "".join(out)).strip("_")
    return base or "teacher"


def _unique_username(base: str) -> str:
    username = base
    i = 2
    while CustomUser.objects.filter(username=username).exists():
        username = f"{base}{i}"
        i += 1
    return username


def _parse_teacher_names_from_upload(uploaded_file) -> list[str]:
    name = (uploaded_file.name or "").lower()
    if name.endswith(".csv"):
        raw = uploaded_file.read()
        text = raw.decode("utf-8-sig", errors="replace")
        rows = []
        for line in text.splitlines():
            if not line.strip():
                continue
            if ";" in line:
                first = line.split(";", 1)[0]
            elif "," in line:
                first = line.split(",", 1)[0]
            else:
                first = line
            rows.append(first.strip())
        return rows

    if name.endswith(".xlsx"):
        try:
            from openpyxl import load_workbook  # type: ignore
        except Exception as exc:
            raise ValueError("Для импорта .xlsx требуется openpyxl") from exc
        buf = io.BytesIO(uploaded_file.read())
        wb = load_workbook(buf, read_only=True, data_only=True)
        ws = wb.active
        rows: list[str] = []
        for row in ws.iter_rows(min_row=1, max_col=1, values_only=True):
            val = row[0]
            if val is None:
                continue
            s = str(val).strip()
            if s:
                rows.append(s)
        return rows

    raise ValueError("Поддерживаются только .csv и .xlsx")


def _target_school_for_school_admin_ops(request):
    user = request.user
    if _is_project_level_user(user):
        school_id = request.session.get("active_school_id")
        if not school_id:
            return None
        return School.objects.filter(id=school_id, status=School.STATUS_APPROVED).first()
    return user.school


def entry_view(request):
    if request.user.is_authenticated:
        if _is_project_level_user(request.user):
            return redirect("accounts:project_hub")
        return redirect("replacements:main-menu")
    return render(request, "entry.html")


def school_register_request_view(request):
    if request.method == "POST":
        form = SchoolRegistrationForm(request.POST)
        if form.is_valid():
            school = form.save(commit=False)
            school.status = School.STATUS_PENDING
            school.approved_at = None
            school.save()
            messages.success(request, "Заявка на регистрацию школы отправлена.")
            return redirect("accounts:entry")
    else:
        form = SchoolRegistrationForm()
    return render(request, "school_register_request.html", {"form": form})


@login_required
@user_passes_test(_is_project_level_user)
def project_hub_view(request):
    schools = School.objects.filter(status=School.STATUS_APPROVED).order_by("name")
    pending = School.objects.filter(status=School.STATUS_PENDING).order_by("-created_at")
    rejected = School.objects.filter(status=School.STATUS_REJECTED).order_by("-created_at")
    active_school = None
    active_school_id = request.session.get("active_school_id")
    if active_school_id:
        active_school = School.objects.filter(id=active_school_id).first()
    return render(
        request,
        "project_hub.html",
        {
            "schools": schools,
            "pending_schools": pending,
            "rejected_schools": rejected,
            "active_school": active_school,
        },
    )


@login_required
@user_passes_test(_is_project_level_user)
@require_POST
def project_select_school_view(request, school_id: int):
    school = get_object_or_404(School, id=school_id, status=School.STATUS_APPROVED)
    request.session["active_school_id"] = school.id
    messages.success(request, f"Выбрана школа: {school.name}")
    return redirect("accounts:project_hub")


@login_required
@user_passes_test(_is_project_level_user)
@require_POST
def school_request_approve_view(request, school_id: int):
    school = get_object_or_404(School, id=school_id)
    school.status = School.STATUS_APPROVED
    school.approved_at = timezone.now()
    school.save(update_fields=["status", "approved_at"])
    messages.success(request, f"Заявка одобрена: {school.name}")
    return redirect("accounts:project_hub")


@login_required
@user_passes_test(_is_project_level_user)
@require_POST
def school_request_reject_view(request, school_id: int):
    school = get_object_or_404(School, id=school_id)
    school.status = School.STATUS_REJECTED
    school.save(update_fields=["status"])
    messages.warning(request, f"Заявка отклонена: {school.name}")
    return redirect("accounts:project_hub")


@login_required
@user_passes_test(_is_project_level_user)
def school_request_review_view(request, school_id: int):
    school = get_object_or_404(School, id=school_id)
    if request.method == "POST":
        form = SchoolReviewForm(request.POST, instance=school)
        if form.is_valid():
            form.save()
            messages.success(request, "Данные школы обновлены.")
            next_action = (request.POST.get("next_action") or "").strip()
            if next_action == "approve":
                school.status = School.STATUS_APPROVED
                school.approved_at = timezone.now()
                school.save(update_fields=["status", "approved_at"])
                messages.success(request, f"Заявка одобрена: {school.name}")
                return redirect("accounts:project_hub")
            if next_action == "reject":
                school.status = School.STATUS_REJECTED
                school.save(update_fields=["status"])
                messages.warning(request, f"Заявка отклонена: {school.name}")
                return redirect("accounts:project_hub")
    else:
        form = SchoolReviewForm(instance=school)
    return render(request, "school_request_review.html", {"school_obj": school, "form": form})


@login_required
@user_passes_test(_is_project_level_user)
def project_users_view(request):
    users = CustomUser.objects.filter(school__isnull=True).order_by("username")
    create_form = ProjectUserCreateForm(request.POST or None)
    if request.method == "POST":
        if create_form.is_valid():
            create_form.save()
            messages.success(request, "Пользователь проекта создан.")
            return redirect("accounts:project_users")
    return render(request, "project_users.html", {"users": users, "create_form": create_form})


@login_required
@user_passes_test(_is_project_level_user)
@require_POST
def project_user_delete_view(request, user_id: int):
    user_obj = get_object_or_404(CustomUser, id=user_id, school__isnull=True)
    if user_obj.is_superuser:
        messages.error(request, "Нельзя удалить суперпользователя.")
        return redirect("accounts:project_users")
    user_obj.delete()
    messages.success(request, "Пользователь проекта удалён.")
    return redirect("accounts:project_users")


@login_required
@user_passes_test(_can_manage_users)
def register_view(request):
    if request.method == "POST":
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            actor = request.user
            if not _is_project_level_user(actor):
                user.school = actor.school
            user.save()
            return redirect("accounts:user_list")
    else:
        form = RegisterForm()
    return render(request, "register.html", {"form": form})


@login_required
def user_list_view(request):
    if not _can_manage_users(request.user):
        return HttpResponse("Forbidden", status=403)
    scope = (request.GET.get("scope") or "").strip().lower()
    qs = CustomUser.objects.all().order_by("username")
    is_project = _is_project_level_user(request.user)

    if is_project:
        if scope == "project":
            qs = qs.filter(school__isnull=True)
        else:
            active_school_id = request.session.get("active_school_id")
            if active_school_id:
                qs = qs.filter(school_id=active_school_id)
            else:
                qs = qs.filter(school__isnull=True)
                scope = "project"
    else:
        qs = qs.filter(school=request.user.school)
        scope = "school"

    return render(
        request,
        "users_list.html",
        {
            "users": qs,
            "scope": scope or "school",
            "is_project_level": is_project,
        },
    )


@login_required
@require_POST
def export_teacher_credentials_view(request):
    if not _can_manage_users(request.user):
        return HttpResponse("Forbidden", status=403)

    qs = CustomUser.objects.filter(is_teacher=True).order_by("full_name", "username")
    if not _is_project_level_user(request.user):
        qs = qs.filter(school=request.user.school)
    users = list(qs)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="teacher_accounts_{ts}.csv"'
    response.write("\ufeff")
    writer = csv.writer(response, delimiter=";")
    writer.writerow(["full_name", "username"])
    for user in users:
        writer.writerow([user.full_name or "", user.username or ""])
    return response


@login_required
@require_POST
def import_teachers_from_file_view(request):
    if not _can_manage_users(request.user):
        return HttpResponse("Forbidden", status=403)

    scope = (request.POST.get("scope") or "").strip().lower()
    if _is_project_level_user(request.user) and scope == "project":
        return HttpResponse("Forbidden", status=403)

    school = _target_school_for_school_admin_ops(request)
    if school is None:
        messages.error(request, "Сначала выберите школу в проектном кабинете.")
        return redirect("accounts:user_list")

    uploaded = request.FILES.get("teachers_file")
    if not uploaded:
        messages.error(request, "Файл не выбран.")
        return redirect("accounts:user_list")

    try:
        names = _parse_teacher_names_from_upload(uploaded)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect("accounts:user_list")

    created_rows: list[list[str]] = []
    skipped_empty = 0
    skipped_exists = 0

    for raw_name in names:
        full_name = (raw_name or "").strip()
        if not full_name:
            skipped_empty += 1
            continue

        if CustomUser.objects.filter(school=school, full_name__iexact=full_name).exists():
            skipped_exists += 1
            continue

        base = _translit_login(full_name)
        username = _unique_username(base)
        password = secrets.token_urlsafe(9)[:12]

        user = CustomUser(
            username=username,
            full_name=full_name,
            school=school,
            is_teacher=True,
            is_admin=False,
            is_guest=False,
            is_system_account=False,
            is_support_system=False,
            is_project_admin=False,
            is_active=True,
        )
        user.set_password(password)
        user.save()

        created_rows.append([full_name, username, password, "Учитель", school.name])

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="created_teacher_accounts_{ts}.csv"'
    response.write("\ufeff")
    writer = csv.writer(response, delimiter=";")
    writer.writerow(["ФИО", "логин", "пароль", "роль", "школа"])
    writer.writerows(created_rows)

    messages.success(
        request,
        f"Создано: {len(created_rows)}. Пропущено пустых: {skipped_empty}. Пропущено дублей: {skipped_exists}.",
    )
    return response


@login_required
def user_edit_view(request, user_id):
    if not _can_manage_users(request.user):
        return HttpResponse("Forbidden", status=403)

    qs = CustomUser.objects.all()
    if not _is_project_level_user(request.user):
        qs = qs.filter(school=request.user.school)
    user_obj = get_object_or_404(qs, id=user_id)
    if getattr(user_obj, "is_admin", False) and not (request.user.is_superuser or getattr(request.user, "is_project_admin", False)):
        return HttpResponse("Forbidden", status=403)
    if request.method == "POST":
        form = UserUpdateForm(request.POST, instance=user_obj, actor=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, "Пользователь успешно обновлен")
            return redirect("accounts:user_list")
    else:
        form = UserUpdateForm(instance=user_obj, actor=request.user)

    return render(request, "user_edit.html", {"form": form, "user_obj": user_obj})


@login_required
@require_POST
def user_delete_view(request, user_id):
    if not _can_manage_users(request.user):
        return HttpResponse("Forbidden", status=403)

    qs = CustomUser.objects.all()
    if not _is_project_level_user(request.user):
        qs = qs.filter(school=request.user.school)
    user_obj = get_object_or_404(qs, id=user_id)
    if getattr(user_obj, "is_admin", False) and not (request.user.is_superuser or getattr(request.user, "is_project_admin", False)):
        return HttpResponse("Forbidden", status=403)
    if request.user == user_obj:
        messages.error(request, "Нельзя удалить текущего пользователя")
        return redirect("accounts:user_list")

    user_obj.delete()
    messages.success(request, "Пользователь удален")
    return redirect("accounts:user_list")


@ratelimit(key="ip", rate="20/m", block=True, method="POST")
def login_view(request):
    if request.method == "POST":
        form = LoginForm(data=request.POST)
        if form.is_valid():
            login(request, form.get_user())
            if _is_project_level_user(request.user):
                return redirect("accounts:project_hub")
            return redirect("replacements:main-menu")
    else:
        form = LoginForm()
    return render(request, "login.html", {"form": form})


@require_POST
def logout_view(request):
    logout(request)
    return redirect("accounts:login")
